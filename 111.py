from tkinter import *
import tkinter.ttk as ttk
import tkinter.font as tkFont  #폰트설정
import tkinter.messagebox as msgbox
import pandas as pd
import openpyxl
from openpyxl.styles.fonts import Font
from datetime import date

root = Tk()
root.title('재셋팅 분리요청건')

# (참고자료:https://passing-story.tistory.com/98)
e = ["", "", "", "", "", "", "", "", "","","","","","","","","","","",""]
label = ["","","","","","","","","","","","","","","","","","","",""]

def input_data() :
    psno = psno_input_txt.get()
    judging_txt = psno[:2]
    psno_int = psno[2:]

    # 1.엑셀파일 만들기#
    def make_excel() :
        # 1) 기본데이터
        serial_data = []
        num = []
        for i in range(0, 20):
            serial_number = e[i]
            serial_data.append(serial_number.get())
            num.append(i)

        selected_comapny = company_name_cmb.get()
        title = selected_comapny + " 재셋팅 제품"  # 제목
        resetting_data = pd.DataFrame({'재셋팅 필요한 제품 S/N': serial_data}, index=num)  # 시리얼정보 데이터
        today = str(date.today())
        tday = today[2:4] + today[5:7] + today[8:]  # 시트명 및 파일명 만들기 : ex)220105
        file_dir = './' + tday + '_' + selected_comapny + '.xlsx'

        # 2) 엑셀파일 만들기
        resetting_data.to_excel(file_dir, sheet_name = tday, startrow=6, startcol= 1 )
        wb = openpyxl.load_workbook(file_dir)
        sheet = wb[tday]
        sheet.cell(2,2).value = title
        sheet['B2'].font = Font(size=16, bold=True, name='HY엽서M')
        sheet.cell(4,2).value = '업체명'
        sheet['B4'].font = Font(size=11, bold=True, name='HY엽서L')
        sheet.cell(4,3).value = selected_comapny
        sheet.cell(5,2).value = '발주번호'
        sheet['B5'].font = Font(size=11, bold=True, name='HY엽서L')
        sheet.cell(5,3).value = psno_input_txt.get()
        sheet.cell(7,2).value = 'No'
        sheet['B7'].font = Font(size=11, bold=True)
        sheet['C7'].font = Font(size=11, bold=True)
        wb.save(file_dir)
        msgbox.showinfo('완료!', '입력완료!')
        print('파일생성 / 데이터입력 완료')

    # 2.입력값이 올바른지 확인
    if psno == "" :
        msgbox.showwarning("발주번호 누락", "발주번호를 입력하세요.")

    elif (len(psno) != 5) and (len(psno) != 7): # 발주번호가 5글자 7글자가 아니라면
        msgbox.showwarning("형식오류", "정확한 발주번호를 입력해주세요1")
    elif len(psno) == 5 : # 발주번호 5글자라면,
        try:
            if type(int(psno)) == int: # 발주번호 5글자가 숫자라면
                psno_input_txt.insert(0, "PS")
                make_excel()
                print('엑셀만들기 실행완료')
        except ValueError :
            msgbox.showwarning('형식오류', "정확한 발주번호를 입력해주세요")
    elif len(psno) == 7 : # 발주번호가 7글자라면
        if (judging_txt == 'ps') or (judging_txt == 'PS') : #앞에가 ps이면
            try:
                if type(int(psno_int)) == int :
                    psno_uppered = judging_txt.upper() + psno_int
                    psno_input_txt.delete(0, END)
                    psno_input_txt.insert(0, psno_uppered)
                    make_excel()
            except ValueError : # 뒤에부분 5글자가 문자라면
                   msgbox.showwarning("형식오류", "정확한 발주번호를 입력해주세요")
        else :
            msgbox.showwarning("형식오류", "정확한 발주번호를 입력해주세요")

########### layout 만들기 #######################

## 1)헤드라인 ##
title_lbl = Label(root, text='재셋팅 분리 제품', padx=5, pady=5)
title_lbl.grid(row=0, column=0, columnspan=2, padx=5, pady=5)

tilt_font = tkFont.Font(family='Arial Black', size=16, weight='bold')
title_lbl.configure(font = tilt_font)

## 2)정보입력 ##

# 업체 정보
company_name_lbl = Label(root, text='업체명')
company_name_lbl.grid(row=1, column=0, sticky='w', padx=5, pady=5)

company_name_opt = ['에이치비테크놀러지', '바오스']
company_name_cmb =  ttk.Combobox(root, state='readonly', values = company_name_opt ,height=2, width=18)
company_name_cmb.current(0)
company_name_cmb.grid(row=1, column=1, padx=5, pady=5)

# 발주번호 정보
psno_input_lbl = Label(root, text = '기존 발주번호')
psno_input_lbl.grid(row=2, column=0, sticky='w', padx=5)

psno_input_txt = Entry(root, width=10)
psno_input_txt.grid(row=2, column=1, sticky='w', padx=5)

# 시리얼 번호 입력
num_lbl = Label(root, text='NO', pady=5)
num_lbl.grid(row=4, column=0, pady=5)

serial_lbl = Label(root, text='재셋팅 필요한 제품 S/N', pady=5)
serial_lbl.grid(row=4, column=1, sticky='w', pady=5)

r = 6
for n in range(0,20) :
    label[n] = Label(root, text=n+1)
    label[n].grid(row=r, column=0, padx=3, pady=1)
    e[n] = Entry(root, width=15)
    e[n].grid(row=r, column=1, sticky='w', padx=3, pady=1)
    r = r + 1

# 버튼
add_btn = Button(root, text='입력 완료', padx=5, pady=5, command = input_data)
add_btn.grid(row=26, column=1, padx=5 ,pady=5)

root.resizable(False, False)
root.mainloop()